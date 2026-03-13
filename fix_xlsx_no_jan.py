import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.chart.data_source import NumVal, StrVal
import sys
sys.stdout.reconfigure(encoding='utf-8')

src = '事業管理/建物明渡/【全保連様】25年7月から26年1月解除案件退去率まとめ【260310】.xlsx'
dst = '事業管理/建物明渡/【全保連様】25年7月から25年12月解除案件退去率まとめ【260310】.xlsx'

# 1月を除外した月別シート
monthly_sheets = ['25年7月解除','25年8月解除','25年9月解除','25年10月解除','25年11月解除','25年12月解除']

# 3件のI列追加
i_col_additions = {
    '25年9月解除': [{'search': '小野', 'date': datetime.datetime(2026, 3, 10)}],
    '25年11月解除': [{'search': '岩﨑', 'date': datetime.datetime(2026, 3, 1)}],
    '25年12月解除': [{'search': '大沢', 'date': datetime.datetime(2026, 3, 1)}],
}

def calc_sheet(file, sheet, additions=None):
    df = pd.read_excel(file, sheet_name=sheet, header=None)
    total = int(df.iloc[1, 1])
    pre_taikyo = 0; post_taikyo = 0; pre_nyukin = 0; post_nyukin = 0
    pre_total = 0; post_total = 0
    taikyo_days = []
    override_map = {}
    if additions:
        for add in additions:
            for idx in range(12, len(df)):
                row_str = ' '.join([str(df.iloc[idx, c]) for c in range(10) if pd.notna(df.iloc[idx, c])])
                if add['search'] in row_str:
                    override_map[idx] = add['date']
                    break
    for idx in range(12, len(df)):
        if pd.isna(df.iloc[idx, 0]) and pd.isna(df.iloc[idx, 1]):
            continue
        sued = isinstance(df.iloc[idx, 6], datetime.datetime)
        if idx in override_map:
            val_i = override_map[idx]; taikyo = True
        else:
            val_i = df.iloc[idx, 8]
            taikyo = isinstance(val_i, datetime.datetime) if pd.notna(val_i) else False
        val_j = df.iloc[idx, 9]
        nyukin = isinstance(val_j, datetime.datetime) if pd.notna(val_j) else False
        kaijo = df.iloc[idx, 4] if isinstance(df.iloc[idx, 4], datetime.datetime) else None
        if sued: post_total += 1
        else: pre_total += 1
        if taikyo:
            if sued: post_taikyo += 1
            else: pre_taikyo += 1
            if kaijo: taikyo_days.append((val_i - kaijo).days)
        if nyukin:
            if sued: post_nyukin += 1
            else: pre_nyukin += 1
    return {
        'total': total, 'pre_total': pre_total, 'post_total': post_total,
        'nini_taikyo': pre_taikyo + post_taikyo,
        'zengaku': pre_nyukin + post_nyukin,
        'kaiketsu': pre_taikyo + post_taikyo + pre_nyukin + post_nyukin,
        'pre_taikyo': pre_taikyo, 'post_taikyo': post_taikyo,
        'pre_nyukin': pre_nyukin, 'post_nyukin': post_nyukin,
        'pre_kaiketsu': pre_taikyo + pre_nyukin,
        'post_kaiketsu': post_taikyo + post_nyukin,
        'kisoritsu': post_total / total if total > 0 else 0,
        'taikyo_days': taikyo_days
    }

# Calculate 7月～12月 only
results = {}
for sheet in monthly_sheets:
    results[sheet] = calc_sheet(src, sheet, i_col_additions.get(sheet))
    r = results[sheet]
    print(f"{sheet}: total={r['total']} taikyo={r['nini_taikyo']} nyukin={r['zengaku']} kaiketsu={r['kaiketsu']}")

# Totals (without 1月)
totals = {'taikyo_days': []}
for key in ['total','pre_total','post_total','nini_taikyo','zengaku','kaiketsu',
            'pre_taikyo','post_taikyo','pre_nyukin','post_nyukin','pre_kaiketsu','post_kaiketsu']:
    totals[key] = sum(r[key] for r in results.values())
for r in results.values():
    totals['taikyo_days'].extend(r['taikyo_days'])
avg_days = sum(totals['taikyo_days']) / len(totals['taikyo_days']) if totals['taikyo_days'] else 0
sorted_days = sorted(totals['taikyo_days'])
median_days = sorted_days[len(sorted_days)//2] if sorted_days else 0
totals['kisoritsu'] = totals['post_total'] / totals['total'] if totals['total'] else 0
t = totals['total']; pre = totals['pre_total']; post = totals['post_total']
print(f"\n合算(7-12月): total={t} taikyo={totals['nini_taikyo']} nyukin={totals['zengaku']} kaiketsu={totals['kaiketsu']}")
print(f"avg_days={avg_days:.1f} median={median_days}")

# Load workbook
wb = load_workbook(src)

# === Remove 26年1月解除 sheet ===
del wb['26年1月解除']
print("Deleted 26年1月解除 sheet")

# === Write 3 I列 dates ===
for sheet, additions in i_col_additions.items():
    ws = wb[sheet]
    df = pd.read_excel(src, sheet_name=sheet, header=None)
    for add in additions:
        for idx in range(12, len(df)):
            row_str = ' '.join([str(df.iloc[idx, c]) for c in range(10) if pd.notna(df.iloc[idx, c])])
            if add['search'] in row_str:
                ws.cell(row=idx+1, column=9).value = add['date']
                print(f"Written I列: {sheet} row {idx+1}")
                break

# === Update monthly sheet summaries ===
def update_summary(ws, r):
    t = r['total']; pre = r['pre_total']; post = r['post_total']
    ws.cell(row=3, column=2).value = r['kaiketsu']
    ws.cell(row=3, column=3).value = r['kaiketsu'] / t if t else 0
    ws.cell(row=4, column=2).value = r['nini_taikyo']
    ws.cell(row=4, column=3).value = r['nini_taikyo'] / t if t else 0
    ws.cell(row=5, column=2).value = r['zengaku']
    ws.cell(row=5, column=3).value = r['zengaku'] / t if t else 0
    ws.cell(row=6, column=2).value = pre
    ws.cell(row=6, column=5).value = post
    ws.cell(row=6, column=7).value = r['kisoritsu']
    ws.cell(row=7, column=2).value = r['pre_kaiketsu']
    ws.cell(row=7, column=3).value = r['pre_kaiketsu'] / pre if pre else 0
    ws.cell(row=7, column=5).value = r['post_kaiketsu']
    ws.cell(row=7, column=6).value = r['post_kaiketsu'] / post if post else 0
    ws.cell(row=8, column=2).value = r['pre_taikyo']
    ws.cell(row=8, column=3).value = r['pre_taikyo'] / pre if pre else 0
    ws.cell(row=8, column=5).value = r['post_taikyo']
    ws.cell(row=8, column=6).value = r['post_taikyo'] / post if post else 0
    ws.cell(row=9, column=2).value = r['pre_nyukin']
    ws.cell(row=9, column=3).value = r['pre_nyukin'] / pre if pre else 0
    ws.cell(row=9, column=5).value = r['post_nyukin']
    ws.cell(row=9, column=6).value = r['post_nyukin'] / post if post else 0

for sheet in monthly_sheets:
    update_summary(wb[sheet], results[sheet])

# === Update 合算 sheet ===
ws_total = wb['25年7月～26年1月合算']
ws_total.title = '25年7月～25年12月合算'
# Update row 1 title text
old_title = ws_total.cell(row=1, column=1).value
if old_title:
    new_title = old_title.replace('2026年1月', '2025年12月').replace('26年1月', '25年12月')
    ws_total.cell(row=1, column=1).value = new_title
# Update row 2 total
ws_total.cell(row=2, column=2).value = t
update_summary(ws_total, totals)
ws_total.cell(row=3, column=5).value = round(avg_days, 1)
ws_total.cell(row=4, column=5).value = median_days
ws_total.cell(row=5, column=5).value = totals['kisoritsu']
# Update row 6 label
r6_label = ws_total.cell(row=6, column=1).value
if r6_label:
    ws_total.cell(row=6, column=1).value = r6_label.replace('26年1月', '25年12月')
print("Updated 合算 -> 25年7月～25年12月合算")

# === Update グラフ分析 ===
ws_g = wb['グラフ分析']

# Row 1 title
old_g_title = ws_g.cell(row=1, column=1).value
if old_g_title:
    ws_g.cell(row=1, column=1).value = old_g_title.replace('2026年1月', '2025年12月').replace('26年1月', '25年12月')

mikaik = t - totals['nini_taikyo'] - totals['zengaku']
mikaik_pre = pre - totals['pre_taikyo'] - totals['pre_nyukin']
mikaik_post = post - totals['post_taikyo'] - totals['post_nyukin']

# ① 全体 (rows 5-7)
ws_g.cell(row=5, column=2).value = totals['nini_taikyo']
ws_g.cell(row=5, column=3).value = totals['nini_taikyo'] / t
ws_g.cell(row=6, column=2).value = totals['zengaku']
ws_g.cell(row=6, column=3).value = totals['zengaku'] / t
ws_g.cell(row=7, column=2).value = mikaik
ws_g.cell(row=7, column=3).value = mikaik / t

# ② 提訴前 (rows 11-13)
ws_g.cell(row=11, column=2).value = totals['pre_taikyo']
ws_g.cell(row=11, column=3).value = totals['pre_taikyo'] / pre
ws_g.cell(row=12, column=2).value = totals['pre_nyukin']
ws_g.cell(row=12, column=3).value = totals['pre_nyukin'] / pre
ws_g.cell(row=13, column=2).value = mikaik_pre
ws_g.cell(row=13, column=3).value = mikaik_pre / pre

# ③ 提訴後 (rows 17-19)
ws_g.cell(row=17, column=2).value = totals['post_taikyo']
ws_g.cell(row=17, column=3).value = totals['post_taikyo'] / post
ws_g.cell(row=18, column=2).value = totals['post_nyukin']
ws_g.cell(row=18, column=3).value = totals['post_nyukin'] / post
ws_g.cell(row=19, column=2).value = mikaik_post
ws_g.cell(row=19, column=3).value = mikaik_post / post

# ④ 月別推移 - 6 months (rows 23-28), clear row 29 (1月)
month_labels = ['7月','8月','9月','10月','11月','12月']
for i, sheet in enumerate(monthly_sheets):
    r = results[sheet]
    row = 23 + i
    ws_g.cell(row=row, column=1).value = month_labels[i]
    ws_g.cell(row=row, column=2).value = r['kaiketsu'] / r['total'] if r['total'] else 0
    ws_g.cell(row=row, column=3).value = r['nini_taikyo'] / r['total'] if r['total'] else 0
    ws_g.cell(row=row, column=4).value = r['total']

# Clear row 29 (was 1月)
for c in range(1, 5):
    ws_g.cell(row=29, column=c).value = None

print("Updated グラフ分析 cells")

# === Update charts ===
charts = ws_g._charts
print(f"Charts count: {len(charts)}")

# Pie charts: update cache values
def update_pie_cache(chart, vals):
    cache = chart.series[0].val.numRef.numCache
    for i, v in enumerate(vals):
        cache.pt[i].v = str(v)

update_pie_cache(charts[0], [totals['nini_taikyo']/t, totals['zengaku']/t, mikaik/t])
update_pie_cache(charts[1], [totals['pre_taikyo']/pre, totals['pre_nyukin']/pre, mikaik_pre/pre])
update_pie_cache(charts[2], [totals['post_taikyo']/post, totals['post_nyukin']/post, mikaik_post/post])

# Update pie chart titles with new counts
def update_chart_title_count(chart, old_count, new_count):
    if old_count == new_count: return
    title_tx = chart.title.tx
    if title_tx and title_tx.rich:
        for para in title_tx.rich.paragraphs:
            for run in para.r:
                if str(old_count) in run.t:
                    run.t = run.t.replace(str(old_count), str(new_count))
                    print(f"  Chart title: {old_count} -> {new_count}")

update_chart_title_count(charts[0], 495, t)     # ① 全体
update_chart_title_count(charts[1], 382, pre)    # ② 提訴前
update_chart_title_count(charts[2], 113, post)   # ③ 提訴後

# ④ Line chart: change data range from 7 months to 6 months
line_chart = charts[3]

# Update series data references from B23:B29 -> B23:B28, C23:C29 -> C23:C28
line_chart.series[0].val.numRef.f = 'グラフ分析!$B$23:$B$28'
line_chart.series[1].val.numRef.f = 'グラフ分析!$C$23:$C$28'

# Update category references from A23:A29 -> A23:A28
line_chart.series[0].cat.strRef.f = 'グラフ分析!$A$23:$A$28'
line_chart.series[1].cat.strRef.f = 'グラフ分析!$A$23:$A$28'

# Update numCache for line chart: 6 points instead of 7
for series_idx in range(2):
    cache = line_chart.series[series_idx].val.numRef.numCache
    # Remove the 7th point (1月) and update ptCount
    cache.pt = cache.pt[:6]
    cache.ptCount = 6
    # Update values
    for i, sheet in enumerate(monthly_sheets):
        r = results[sheet]
        if series_idx == 0:  # 解決率
            cache.pt[i].v = str(r['kaiketsu'] / r['total'] if r['total'] else 0)
        else:  # 任意退去率
            cache.pt[i].v = str(r['nini_taikyo'] / r['total'] if r['total'] else 0)

# Update category cache for line chart
for series_idx in range(2):
    cat_cache = line_chart.series[series_idx].cat.strRef.strCache
    cat_cache.pt = cat_cache.pt[:6]
    cat_cache.ptCount = 6

# Update line chart title
title_tx = line_chart.title.tx
if title_tx and title_tx.rich:
    for para in title_tx.rich.paragraphs:
        for run in para.r:
            # No specific count in title ④, just the description
            pass

print("Updated all charts")

wb.save(dst)
print(f"\nSaved: {dst}")
