import pandas as pd
import datetime
from openpyxl import load_workbook
from copy import copy
import sys
sys.stdout.reconfigure(encoding='utf-8')

src = '事業管理/建物明渡/【全保連様】25年7月から26年1月解除案件退去率まとめ【260310】.xlsx'
dst = '事業管理/建物明渡/【全保連様】25年7月から26年1月解除案件退去率まとめ【260310】_修正版v3.xlsx'

monthly_sheets = ['25年7月解除','25年8月解除','25年9月解除','25年10月解除','25年11月解除','25年12月解除','26年1月解除']

# === Step 0: Define the 3 I列 additions from CSV ===
i_col_additions = {
    '25年9月解除': [{'search': '小野', 'date': datetime.datetime(2026, 3, 10)}],
    '25年11月解除': [{'search': '岩﨑', 'date': datetime.datetime(2026, 3, 1)}],
    '25年12月解除': [{'search': '大沢', 'date': datetime.datetime(2026, 3, 1)}],
}

def calc_sheet(file, sheet, additions=None):
    df = pd.read_excel(file, sheet_name=sheet, header=None)
    total = int(df.iloc[1, 1])
    pre_taikyo = 0; post_taikyo = 0; pre_nyukin = 0; post_nyukin = 0
    pre_total = 0; post_total = 0
    taikyo_days = []

    # Build override map: row_idx -> date for I列 additions
    override_map = {}
    if additions:
        for add in additions:
            for idx in range(12, len(df)):
                row_str = ' '.join([str(df.iloc[idx, c]) for c in range(10) if pd.notna(df.iloc[idx, c])])
                if add['search'] in row_str:
                    override_map[idx] = add['date']
                    print(f"  I列追加: {sheet} row {idx+1} ({add['search']}) -> {add['date'].strftime('%Y-%m-%d')}")
                    break

    for idx in range(12, len(df)):
        if pd.isna(df.iloc[idx, 0]) and pd.isna(df.iloc[idx, 1]):
            continue
        sued = isinstance(df.iloc[idx, 6], datetime.datetime)

        # Check I列: use override if present, else original
        if idx in override_map:
            val_i = override_map[idx]
            taikyo = True
        else:
            val_i = df.iloc[idx, 8]
            taikyo = isinstance(val_i, datetime.datetime) if pd.notna(val_i) else False

        val_j = df.iloc[idx, 9]
        nyukin = isinstance(val_j, datetime.datetime) if pd.notna(val_j) else False
        kaijo = df.iloc[idx, 4] if isinstance(df.iloc[idx, 4], datetime.datetime) else None

        if sued:
            post_total += 1
        else:
            pre_total += 1
        if taikyo:
            if sued: post_taikyo += 1
            else: pre_taikyo += 1
            if kaijo:
                taikyo_days.append((val_i - kaijo).days)
        if nyukin:
            if sued: post_nyukin += 1
            else: pre_nyukin += 1

    return {
        'total': total, 'pre_total': pre_total, 'post_total': post_total,
        'nini_taikyo': pre_taikyo + post_taikyo,
        'zengaku': pre_nyukin + post_nyukin,
        'kaiketsu': pre_taikyo + post_taikyo + pre_nyukin + post_nyukin,
        'pre_taikyo': pre_taikyo, 'post_taikyo': post_taikyo,
        'pre_nyukin': pre_nyukin, 'post_nyukin': post_nyukin,
        'pre_kaiketsu': pre_taikyo + pre_nyukin,
        'post_kaiketsu': post_taikyo + post_nyukin,
        'kisoritsu': post_total / total if total > 0 else 0,
        'taikyo_days': taikyo_days
    }

# Calculate all months (with additions applied)
results = {}
for sheet in monthly_sheets:
    additions = i_col_additions.get(sheet, None)
    results[sheet] = calc_sheet(src, sheet, additions)
    r = results[sheet]
    print(f"{sheet}: total={r['total']} taikyo={r['nini_taikyo']} nyukin={r['zengaku']} kaiketsu={r['kaiketsu']}")

# Calculate totals
totals = {'taikyo_days': []}
for key in ['total','pre_total','post_total','nini_taikyo','zengaku','kaiketsu',
            'pre_taikyo','post_taikyo','pre_nyukin','post_nyukin','pre_kaiketsu','post_kaiketsu']:
    totals[key] = sum(r[key] for r in results.values())
for r in results.values():
    totals['taikyo_days'].extend(r['taikyo_days'])
avg_days = sum(totals['taikyo_days']) / len(totals['taikyo_days']) if totals['taikyo_days'] else 0
sorted_days = sorted(totals['taikyo_days'])
median_days = sorted_days[len(sorted_days)//2] if sorted_days else 0
totals['kisoritsu'] = totals['post_total'] / totals['total'] if totals['total'] else 0
print(f"\n合算: taikyo={totals['nini_taikyo']} nyukin={totals['zengaku']} kaiketsu={totals['kaiketsu']}")
print(f"avg_days={avg_days:.1f} median={median_days}")

# Load workbook - preserving everything including charts
wb = load_workbook(src)

# === Write the 3 I列 dates into the workbook ===
for sheet, additions in i_col_additions.items():
    ws = wb[sheet]
    df = pd.read_excel(src, sheet_name=sheet, header=None)
    for add in additions:
        for idx in range(12, len(df)):
            row_str = ' '.join([str(df.iloc[idx, c]) for c in range(10) if pd.notna(df.iloc[idx, c])])
            if add['search'] in row_str:
                ws.cell(row=idx+1, column=9).value = add['date']
                print(f"Written I列: {sheet} row {idx+1} = {add['date']}")
                break

def update_summary(ws, r):
    t = r['total']; pre = r['pre_total']; post = r['post_total']
    ws.cell(row=3, column=2).value = r['kaiketsu']
    ws.cell(row=3, column=3).value = r['kaiketsu'] / t if t else 0
    ws.cell(row=4, column=2).value = r['nini_taikyo']
    ws.cell(row=4, column=3).value = r['nini_taikyo'] / t if t else 0
    ws.cell(row=5, column=2).value = r['zengaku']
    ws.cell(row=5, column=3).value = r['zengaku'] / t if t else 0
    ws.cell(row=6, column=2).value = pre
    ws.cell(row=6, column=5).value = post
    ws.cell(row=6, column=7).value = r['kisoritsu']
    ws.cell(row=7, column=2).value = r['pre_kaiketsu']
    ws.cell(row=7, column=3).value = r['pre_kaiketsu'] / pre if pre else 0
    ws.cell(row=7, column=5).value = r['post_kaiketsu']
    ws.cell(row=7, column=6).value = r['post_kaiketsu'] / post if post else 0
    ws.cell(row=8, column=2).value = r['pre_taikyo']
    ws.cell(row=8, column=3).value = r['pre_taikyo'] / pre if pre else 0
    ws.cell(row=8, column=5).value = r['post_taikyo']
    ws.cell(row=8, column=6).value = r['post_taikyo'] / post if post else 0
    ws.cell(row=9, column=2).value = r['pre_nyukin']
    ws.cell(row=9, column=3).value = r['pre_nyukin'] / pre if pre else 0
    ws.cell(row=9, column=5).value = r['post_nyukin']
    ws.cell(row=9, column=6).value = r['post_nyukin'] / post if post else 0

# Update monthly sheets
for sheet in monthly_sheets:
    update_summary(wb[sheet], results[sheet])
    print(f"Updated {sheet}")

# Update 合算
ws_total = wb['25年7月～26年1月合算']
update_summary(ws_total, totals)
ws_total.cell(row=3, column=5).value = round(avg_days, 1)
ws_total.cell(row=4, column=5).value = median_days
ws_total.cell(row=5, column=5).value = totals['kisoritsu']
print("Updated 合算")

# Update グラフ分析 cell values (charts reference these cells)
ws_g = wb['グラフ分析']
t = totals['total']; pre = totals['pre_total']; post = totals['post_total']

# ① 全体 (rows 5-7)
mikaik = t - totals['nini_taikyo'] - totals['zengaku']
ws_g.cell(row=5, column=2).value = totals['nini_taikyo']
ws_g.cell(row=5, column=3).value = totals['nini_taikyo'] / t
ws_g.cell(row=6, column=2).value = totals['zengaku']
ws_g.cell(row=6, column=3).value = totals['zengaku'] / t
ws_g.cell(row=7, column=2).value = mikaik
ws_g.cell(row=7, column=3).value = mikaik / t

# ② 提訴前 (rows 11-13)
mikaik_pre = pre - totals['pre_taikyo'] - totals['pre_nyukin']
ws_g.cell(row=11, column=2).value = totals['pre_taikyo']
ws_g.cell(row=11, column=3).value = totals['pre_taikyo'] / pre
ws_g.cell(row=12, column=2).value = totals['pre_nyukin']
ws_g.cell(row=12, column=3).value = totals['pre_nyukin'] / pre
ws_g.cell(row=13, column=2).value = mikaik_pre
ws_g.cell(row=13, column=3).value = mikaik_pre / pre

# ③ 提訴後 (rows 17-19)
mikaik_post = post - totals['post_taikyo'] - totals['post_nyukin']
ws_g.cell(row=17, column=2).value = totals['post_taikyo']
ws_g.cell(row=17, column=3).value = totals['post_taikyo'] / post
ws_g.cell(row=18, column=2).value = totals['post_nyukin']
ws_g.cell(row=18, column=3).value = totals['post_nyukin'] / post
ws_g.cell(row=19, column=2).value = mikaik_post
ws_g.cell(row=19, column=3).value = mikaik_post / post

# ④ 月別推移 (rows 23-29)
for i, sheet in enumerate(monthly_sheets):
    r = results[sheet]
    row = 23 + i
    ws_g.cell(row=row, column=2).value = r['kaiketsu'] / r['total'] if r['total'] else 0
    ws_g.cell(row=row, column=3).value = r['nini_taikyo'] / r['total'] if r['total'] else 0
    ws_g.cell(row=row, column=4).value = r['total']

print("Updated グラフ分析 cells")

# Now update chart numCache so charts show correct values without needing Excel recalc
charts = ws_g._charts
print(f"Charts preserved: {len(charts)}")

def update_pie_cache(chart, vals):
    """Update the numCache of a pie chart's series"""
    cache = chart.series[0].val.numRef.numCache
    for i, v in enumerate(vals):
        cache.pt[i].v = str(v)

# ① 全体
v1 = [totals['nini_taikyo']/t, totals['zengaku']/t, mikaik/t]
update_pie_cache(charts[0], v1)

# ② 提訴前
v2 = [totals['pre_taikyo']/pre, totals['pre_nyukin']/pre, mikaik_pre/pre]
update_pie_cache(charts[1], v2)

# ③ 提訴後
v3 = [totals['post_taikyo']/post, totals['post_nyukin']/post, mikaik_post/post]
update_pie_cache(charts[2], v3)

# ④ 月別推移 - series 0 (解決率), series 1 (任意退去率)
line_chart = charts[3]
for i, sheet in enumerate(monthly_sheets):
    r = results[sheet]
    kaiketsu_rate = r['kaiketsu'] / r['total'] if r['total'] else 0
    taikyo_rate = r['nini_taikyo'] / r['total'] if r['total'] else 0
    line_chart.series[0].val.numRef.numCache.pt[i].v = str(kaiketsu_rate)
    line_chart.series[1].val.numRef.numCache.pt[i].v = str(taikyo_rate)

print("Updated chart caches")

# Update chart titles with new counts
def update_chart_title_count(chart, old_count, new_count):
    if old_count == new_count:
        return
    title_tx = chart.title.tx
    if title_tx and title_tx.rich:
        for para in title_tx.rich.paragraphs:
            for run in para.r:
                if str(old_count) in run.t:
                    run.t = run.t.replace(str(old_count), str(new_count))
                    print(f"  Title updated: {old_count} -> {new_count}")

update_chart_title_count(charts[1], 382, totals['pre_total'])
update_chart_title_count(charts[2], 113, totals['post_total'])

wb.save(dst)
print(f"\nSaved: {dst}")
print("Done - charts preserved, I列 3件追加済み")
