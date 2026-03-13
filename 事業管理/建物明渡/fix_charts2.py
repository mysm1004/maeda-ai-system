# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText, Text
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties,
    Font as DrawingFont, RichTextProperties, RegularTextRun
)
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.title import Title
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

FILE = r'C:\Users\user\Desktop\claudeマスター\事業管理\建物明渡\全保連解除＆退去案件まとめ【260309】.xlsx'

wb = load_workbook(FILE)

# グラフ分析シートを削除して作り直す
if 'グラフ分析' in wb.sheetnames:
    del wb['グラフ分析']

ws = wb.create_sheet('グラフ分析')

# ========== 共通スタイル ==========
PIE_COLORS = ['92D050', '5B9BD5', 'BFBFBF']  # 緑, 青, グレー
FONT_HEADER = Font(name='游ゴシック', size=11, bold=True)
FONT_DATA = Font(name='游ゴシック', size=10)
FONT_TITLE = Font(name='游ゴシック', size=12, bold=True)

def make_jp_font(sz=1000, bold=False):
    return CharacterProperties(
        latin=DrawingFont(typeface='游ゴシック'),
        ea=DrawingFont(typeface='游ゴシック'),
        sz=sz, b=bold
    )

def make_chart_title(text, sz=1200):
    run = RegularTextRun(rPr=make_jp_font(sz=sz, bold=True), t=text)
    para = Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=sz, bold=True)), r=[run])
    rich = RichText(p=[para])
    tx = Text(rich=rich)
    return Title(tx=tx)

# ========== 合算シートからデータ再集計 ==========
ws_data = wb['25年7月～26年1月合算']

# 全体集計
total = 0
taijo_total = 0
kaishuu_total = 0
sued_total = 0

# 提訴前
pre_total = 0
pre_taijo = 0
pre_kaishuu = 0

# 提訴後
post_total = 0
post_taijo = 0
post_kaishuu = 0

for row in range(13, ws_data.max_row + 1):
    a_val = ws_data.cell(row=row, column=1).value
    if a_val is None:
        break
    total += 1

    g_val = ws_data.cell(row=row, column=7).value   # 訴訟提起日
    i_val = ws_data.cell(row=row, column=9).value   # 任意退去日
    j_val = ws_data.cell(row=row, column=10).value  # 全額入金日

    has_lawsuit = g_val is not None
    has_taijo = i_val is not None
    has_kaishuu = j_val is not None

    if has_lawsuit:
        sued_total += 1
        post_total += 1
        if has_taijo:
            post_taijo += 1
            taijo_total += 1
        elif has_kaishuu:
            post_kaishuu += 1
            kaishuu_total += 1
    else:
        pre_total += 1
        if has_taijo:
            pre_taijo += 1
            taijo_total += 1
        elif has_kaishuu:
            pre_kaishuu += 1
            kaishuu_total += 1

pre_mikaiketsu = pre_total - pre_taijo - pre_kaishuu
post_mikaiketsu = post_total - post_taijo - post_kaishuu
mikaiketsu_total = total - taijo_total - kaishuu_total

print(f'全体: {total}件, 退去{taijo_total}, 回収{kaishuu_total}, 未解決{mikaiketsu_total}')
print(f'提訴前: {pre_total}件, 退去{pre_taijo}, 回収{pre_kaishuu}, 未解決{pre_mikaiketsu}')
print(f'提訴後: {post_total}件, 退去{post_taijo}, 回収{post_kaishuu}, 未解決{post_mikaiketsu}')

# ========== 月別データ集計 ==========
month_sheets = ['25年7月解除','25年8月解除','25年9月解除','25年10月解除',
                '25年11月解除','25年12月解除','26年1月解除']
month_labels = ['7月','8月','9月','10月','11月','12月','1月']

monthly_data = []
for sheet_name in month_sheets:
    ms = wb[sheet_name]
    m_total = 0
    m_taijo = 0
    m_kaishuu = 0
    for row in range(13, ms.max_row + 1):
        a_val = ms.cell(row=row, column=1).value
        if a_val is None:
            break
        m_total += 1
        i_val = ms.cell(row=row, column=9).value
        j_val = ms.cell(row=row, column=10).value
        if i_val is not None:
            m_taijo += 1
        elif j_val is not None:
            m_kaishuu += 1
    kaiketsu = m_taijo + m_kaishuu
    kaiketsu_rate = kaiketsu / m_total if m_total > 0 else 0
    taijo_rate = m_taijo / m_total if m_total > 0 else 0
    monthly_data.append((m_total, kaiketsu, m_taijo, kaiketsu_rate, taijo_rate))

# ========== シートにデータ書き込み ==========

# 全体タイトル
ws['A1'] = '【2025年7月～2026年1月 解除通知発送案件】グラフ分析（26年3月9日時点）'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

# ===== ① 全体の円グラフデータ =====
ws['A3'] = '① 全体（提訴前＋提訴後合算）の解決状況'
ws['A3'].font = FONT_HEADER

ws['A4'] = 'カテゴリ'
ws['B4'] = '件数'
ws['C4'] = '割合'
for c in ['A4','B4','C4']:
    ws[c].font = FONT_HEADER

# データ：割合を直接計算してセルに入れる
pie1_data = [
    ('任意退去', taijo_total, taijo_total / total),
    ('全額入金（回収）', kaishuu_total, kaishuu_total / total),
    ('未解決', mikaiketsu_total, mikaiketsu_total / total),
]
for i, (cat, cnt, pct) in enumerate(pie1_data):
    r = 5 + i
    ws.cell(row=r, column=1, value=cat).font = FONT_DATA
    ws.cell(row=r, column=2, value=cnt).font = FONT_DATA
    cell_pct = ws.cell(row=r, column=3, value=pct)
    cell_pct.font = FONT_DATA
    cell_pct.number_format = '0.0%'

# ===== ② 提訴前 =====
ws['A10'] = '② 提訴前案件の解決状況'
ws['A10'].font = FONT_HEADER
ws['A11'] = 'カテゴリ'
ws['B11'] = '件数'
ws['C11'] = '割合'
for c in ['A11','B11','C11']:
    ws[c].font = FONT_HEADER

pie2_data = [
    ('任意退去', pre_taijo, pre_taijo / pre_total),
    ('全額入金（回収）', pre_kaishuu, pre_kaishuu / pre_total),
    ('未解決', pre_mikaiketsu, pre_mikaiketsu / pre_total),
]
for i, (cat, cnt, pct) in enumerate(pie2_data):
    r = 12 + i
    ws.cell(row=r, column=1, value=cat).font = FONT_DATA
    ws.cell(row=r, column=2, value=cnt).font = FONT_DATA
    cell_pct = ws.cell(row=r, column=3, value=pct)
    cell_pct.font = FONT_DATA
    cell_pct.number_format = '0.0%'

# ===== ③ 提訴後 =====
ws['A17'] = '③ 提訴後案件の解決状況'
ws['A17'].font = FONT_HEADER
ws['A18'] = 'カテゴリ'
ws['B18'] = '件数'
ws['C18'] = '割合'
for c in ['A18','B18','C18']:
    ws[c].font = FONT_HEADER

pie3_data = [
    ('任意退去', post_taijo, post_taijo / post_total),
    ('全額入金（回収）', post_kaishuu, post_kaishuu / post_total),
    ('未解決', post_mikaiketsu, post_mikaiketsu / post_total),
]
for i, (cat, cnt, pct) in enumerate(pie3_data):
    r = 19 + i
    ws.cell(row=r, column=1, value=cat).font = FONT_DATA
    ws.cell(row=r, column=2, value=cnt).font = FONT_DATA
    cell_pct = ws.cell(row=r, column=3, value=pct)
    cell_pct.font = FONT_DATA
    cell_pct.number_format = '0.0%'

# ===== ④ 月別推移データ =====
ws['A24'] = '④ 月別推移：解決率・任意退去率（提訴前＋提訴後合算）'
ws['A24'].font = FONT_HEADER
ws['A25'] = '月'
ws['B25'] = '解決率'
ws['C25'] = '任意退去率'
for c in ['A25','B25','C25']:
    ws[c].font = FONT_HEADER

for i, (m_total, m_kaiketsu, m_taijo, k_rate, t_rate) in enumerate(monthly_data):
    r = 26 + i
    ws.cell(row=r, column=1, value=month_labels[i]).font = FONT_DATA
    cell_k = ws.cell(row=r, column=2, value=k_rate)
    cell_k.font = FONT_DATA
    cell_k.number_format = '0.0%'
    cell_t = ws.cell(row=r, column=3, value=t_rate)
    cell_t.font = FONT_DATA
    cell_t.number_format = '0.0%'

# 列幅調整
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12

# ========== グラフ作成 ==========

def create_pie(ws, title_text, cat_rows, data_col_pct, anchor, total_n):
    """
    円グラフ: 割合(C列)の値でスライスを作り、
    ラベルには「カテゴリ名: XX.X%（N件）」形式で表示
    """
    pie = PieChart()
    pie.title = make_chart_title(title_text, sz=1100)
    pie.style = 10

    # 割合の列をデータとして使用（C列）
    data_ref = Reference(ws, min_col=data_col_pct, min_row=cat_rows[0], max_row=cat_rows[-1])
    cats_ref = Reference(ws, min_col=1, min_row=cat_rows[0], max_row=cat_rows[-1])
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(cats_ref)

    # 色設定
    series = pie.series[0]
    for i, color in enumerate(PIE_COLORS):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)

    # データラベル: カテゴリ名 + 値（割合）を表示
    # 値は割合のセル(0.251...)をnumFmt='0.0%'で「25.1%」と表示
    series.dLbls = DataLabelList()
    series.dLbls.showPercent = False  # Excel自動計算パーセントは使わない
    series.dLbls.showCatName = True   # カテゴリ名を表示
    series.dLbls.showVal = True       # 値（＝割合セルの値）を表示
    series.dLbls.showSerName = False
    series.dLbls.showLeaderLines = True
    series.dLbls.numFmt = '0.0%'     # 値の表示形式: 0.251 → 25.1%
    series.dLbls.separator = '\n'     # 改行で区切る

    # ラベルフォント
    series.dLbls.txPr = RichText(
        p=[Paragraph(
            pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)),
            r=[]
        )]
    )

    # 凡例非表示（ラベルにカテゴリ名が出るので不要）
    pie.legend = None

    pie.width = 16
    pie.height = 13

    ws.add_chart(pie, anchor)

# ①全体
create_pie(ws, f'① 全体の解決状況（全{total}件）',
           [5, 6, 7], 3, 'E3', total)

# ②提訴前
create_pie(ws, f'② 提訴前の解決状況（全{pre_total}件）',
           [12, 13, 14], 3, 'E10', pre_total)

# ③提訴後
create_pie(ws, f'③ 提訴後の解決状況（全{post_total}件）',
           [19, 20, 21], 3, 'E17', post_total)

# ===== ④ 折れ線グラフ =====
line = LineChart()
line.title = make_chart_title('④ 月別 解決率・任意退去率の推移', sz=1100)
line.style = 10
line.width = 22
line.height = 14

# データ
data_ref = Reference(ws, min_col=2, min_row=25, max_col=3, max_row=32)
cats_ref = Reference(ws, min_col=1, min_row=26, max_row=32)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)

# Y軸
line.y_axis.title = '割合'
line.y_axis.numFmt = '0%'
line.y_axis.scaling.min = 0
line.y_axis.scaling.max = 0.6
line.y_axis.majorUnit = 0.1
line.y_axis.delete = False
line.y_axis.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=900)), r=[])]
)

# X軸
line.x_axis.delete = False
line.x_axis.tickLblPos = 'low'
line.x_axis.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)), r=[])]
)

# 系列スタイル
# 解決率: 青
s0 = line.series[0]
s0.graphicalProperties.line.solidFill = '4472C4'
s0.graphicalProperties.line.width = 28000
s0.marker.symbol = 'circle'
s0.marker.size = 8
s0.marker.graphicalProperties.solidFill = '4472C4'
s0.dLbls = None  # 線上のデータラベル削除

# 任意退去率: 緑
s1 = line.series[1]
s1.graphicalProperties.line.solidFill = '92D050'
s1.graphicalProperties.line.width = 28000
s1.marker.symbol = 'square'
s1.marker.size = 8
s1.marker.graphicalProperties.solidFill = '92D050'
s1.dLbls = None  # 線上のデータラベル削除

# 凡例
line.legend.position = 'b'
line.legend.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)), r=[])]
)

ws.add_chart(line, 'E24')

# ========== 保存 ==========
wb.save(FILE)
print()
print('グラフ分析シート作り直し完了！')
print('改善点:')
print('  円グラフ: 割合の値を直接セルに入れ、numFmt=0.0%で「25.1%」と正確に表示')
print('  円グラフ: カテゴリ名 + 改行 + パーセントのシンプルなラベル')
print('  折れ線: データラベル無し、Y軸0%~60%(10%刻み)、X軸に月名')
