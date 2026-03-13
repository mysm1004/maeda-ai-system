# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = r'C:\Users\user\Desktop\claudeマスター\事業管理\建物明渡\全保連解除＆退去案件まとめ【260309】.xlsx'
wb = load_workbook(FILE)

# ============================================================
# 改善①: 合算シートD3-D4に平均日数を追加
# ============================================================
ws = wb['25年7月～26年1月合算']

# E列→I列の日数計算（日付型のもののみ）
days_list = []
for row in range(13, ws.max_row + 1):
    if ws.cell(row=row, column=1).value is None:
        break
    e_val = ws.cell(row=row, column=5).value  # 解除通知発送日
    i_val = ws.cell(row=row, column=9).value  # 任意退去日
    if i_val is not None and e_val is not None:
        if isinstance(e_val, datetime) and isinstance(i_val, datetime):
            diff = (i_val - e_val).days
            days_list.append(diff)

avg_days = sum(days_list) / len(days_list) if days_list else 0
median_days = sorted(days_list)[len(days_list) // 2] if days_list else 0

print(f'平均日数: {avg_days:.1f}日 (対象{len(days_list)}件, 中央値{median_days}日)')

# D3: ラベル
FONT_8B = Font(name='游ゴシック', size=8, bold=True)
FONT_8 = Font(name='游ゴシック', size=8)

# 既存の行3,4のピンク背景色を取得（他セルと統一）
pink_fill = copy(ws.cell(2, 1).fill)  # A2のfillをコピー
border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

ws.cell(3, 4).value = '解除通知送付～任意退去の平均日数'
ws.cell(3, 4).font = FONT_8B
ws.cell(3, 4).fill = copy(pink_fill)
ws.cell(3, 4).border = copy(border_thin)
ws.cell(3, 4).alignment = Alignment(wrap_text=True, vertical='center')

ws.cell(3, 5).value = round(avg_days, 1)
ws.cell(3, 5).font = FONT_8B
ws.cell(3, 5).fill = copy(pink_fill)
ws.cell(3, 5).border = copy(border_thin)
ws.cell(3, 5).number_format = '0.0"日"'
ws.cell(3, 5).alignment = Alignment(horizontal='center', vertical='center')

ws.cell(4, 4).value = '（中央値）'
ws.cell(4, 4).font = FONT_8B
ws.cell(4, 4).fill = copy(pink_fill)
ws.cell(4, 4).border = copy(border_thin)
ws.cell(4, 4).alignment = Alignment(horizontal='right', vertical='center')

ws.cell(4, 5).value = median_days
ws.cell(4, 5).font = FONT_8B
ws.cell(4, 5).fill = copy(pink_fill)
ws.cell(4, 5).border = copy(border_thin)
ws.cell(4, 5).number_format = '0"日"'
ws.cell(4, 5).alignment = Alignment(horizontal='center', vertical='center')

# F3, F4にも罫線とfillを統一（空欄でも体裁を揃える）
for r in [3, 4]:
    ws.cell(r, 6).fill = copy(pink_fill)
    ws.cell(r, 6).border = copy(border_thin)

print('改善①完了: D3-E4に平均日数・中央値を追加')

# ============================================================
# 改善②: グラフ分析シートの④付近にコメントを追加
# ============================================================
ws2 = wb['グラフ分析']

# コメントをセルに書き込む（④のチャートはE62配置、データはA22-D29）
# チャート下あたり（row 34〜）にコメントエリアを作成
comment_start_row = 34

# コメント背景: 薄い黄色
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FONT_COMMENT_TITLE = Font(name='游ゴシック', size=10, bold=True, color='333333')
FONT_COMMENT = Font(name='游ゴシック', size=9, color='333333')
BORDER_COMMENT = Border(
    left=Side(style='medium', color='E6B800'),
    right=Side(style='medium', color='E6B800'),
    top=Side(style='medium', color='E6B800'),
    bottom=Side(style='medium', color='E6B800'),
)
BORDER_COMMENT_TOP = Border(
    left=Side(style='medium', color='E6B800'),
    right=Side(style='medium', color='E6B800'),
    top=Side(style='medium', color='E6B800'),
)
BORDER_COMMENT_MID = Border(
    left=Side(style='medium', color='E6B800'),
    right=Side(style='medium', color='E6B800'),
)
BORDER_COMMENT_BTM = Border(
    left=Side(style='medium', color='E6B800'),
    right=Side(style='medium', color='E6B800'),
    bottom=Side(style='medium', color='E6B800'),
)

# コメント内容（各行）
comment_lines = [
    ('💡 ④グラフの読み方（分析上の留意点）', FONT_COMMENT_TITLE, BORDER_COMMENT_TOP),
    ('', FONT_COMMENT, BORDER_COMMENT_MID),
    ('① 解除通知送付後から解決（退去）までは一定のタイムラグがある', FONT_COMMENT, BORDER_COMMENT_MID),
    (f'　　（平均 約{round(avg_days)}日 ＝ 約{round(avg_days/30, 1)}ヶ月）', FONT_COMMENT, BORDER_COMMENT_MID),
    ('', FONT_COMMENT, BORDER_COMMENT_MID),
    ('② 解除件数が右肩上がりに増えている', FONT_COMMENT, BORDER_COMMENT_MID),
    ('', FONT_COMMENT, BORDER_COMMENT_MID),
    ('この２点を前提にすると、', FONT_COMMENT, BORDER_COMMENT_MID),
    ('解除通知送付件数に占める退去「率」は一時的に低下する。', FONT_COMMENT, BORDER_COMMENT_MID),
    ('', FONT_COMMENT, BORDER_COMMENT_MID),
    ('➡ この点を踏まえてなお、退去率・解決率の数字が', FONT_COMMENT, BORDER_COMMENT_MID),
    ('　  ○なのか否かが検証対象となる。', FONT_COMMENT, BORDER_COMMENT_BTM),
]

# A34～A45あたりにコメント書き込み（A-D列を結合）
for i, (text, font, border) in enumerate(comment_lines):
    r = comment_start_row + i
    # A列にテキスト
    cell = ws2.cell(r, 1, text)
    cell.font = font
    cell.fill = YELLOW_FILL
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    # A-D列に黄色背景+罫線を展開
    for c in range(1, 5):
        ws2.cell(r, c).fill = YELLOW_FILL
        ws2.cell(r, c).border = border
    # A列結合（A-D）
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

print('改善②完了: グラフ分析シートにコメント追加（黄色背景の囲み枠）')

# ============================================================
# 保存
# ============================================================
wb.save(FILE)
print('\n全改善完了！')
