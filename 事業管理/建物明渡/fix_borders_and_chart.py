# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from copy import copy
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties,
    Font as DrawingFont, RegularTextRun
)
from openpyxl.styles import Font

FILE = r'C:\Users\user\Desktop\claudeマスター\事業管理\建物明渡\全保連解除＆退去案件まとめ【260309】.xlsx'
wb = load_workbook(FILE)

# ============================================================
# タスク①: 7月解除の1-9行目の罫線を8月以降+合算シートに展開
# ============================================================
src = wb['25年7月解除']

target_sheets = [
    '25年8月解除', '25年9月解除', '25年10月解除',
    '25年11月解除', '25年12月解除', '26年1月解除',
    '26年2月解除', '25年7月～26年1月合算'
]

for sheet_name in target_sheets:
    tgt = wb[sheet_name]
    # 1行目～9行目、A～G列(1～7)の罫線をコピー
    for row in range(1, 10):
        for col in range(1, 8):
            src_cell = src.cell(row=row, column=col)
            tgt_cell = tgt.cell(row=row, column=col)
            # 罫線をコピー
            tgt_cell.border = copy(src_cell.border)
    print(f'  罫線コピー完了: {sheet_name}')

print('タスク①完了: 罫線を全シートに展開')

# ============================================================
# タスク②: グラフ分析シートの④折れ線に「解除件数」を追加
# ============================================================
ws = wb['グラフ分析']

# 月別の解除件数を集計
month_sheets = ['25年7月解除','25年8月解除','25年9月解除','25年10月解除',
                '25年11月解除','25年12月解除','26年1月解除']
month_counts = []
for sn in month_sheets:
    ms = wb[sn]
    cnt = 0
    for row in range(13, ms.max_row + 1):
        if ms.cell(row=row, column=1).value is None:
            break
        cnt += 1
    month_counts.append(cnt)

print(f'\n月別解除件数: {month_counts}')

# D列に「解除件数」を追加
ws.cell(22, 4, '解除件数').font = Font(name='游ゴシック', size=11, bold=True)
for i, cnt in enumerate(month_counts):
    ws.cell(23 + i, 4, cnt).font = Font(name='游ゴシック', size=10)

# 既存の④折れ線グラフを削除して作り直す
# グラフは4つ目（index 3）が折れ線
# 全チャートから折れ線を見つけて削除
new_charts = []
line_anchor = 'E62'  # デフォルト
for chart in ws._charts:
    if hasattr(chart, 'y_axis') and isinstance(chart, LineChart):
        # 折れ線グラフをスキップ（削除）
        continue
    new_charts.append(chart)
ws._charts = new_charts

# ---- 新しい折れ線グラフ（2軸）を作成 ----
def make_jp_font(sz=1000, bold=False):
    return CharacterProperties(
        latin=DrawingFont(typeface='游ゴシック'),
        ea=DrawingFont(typeface='游ゴシック'),
        sz=sz, b=bold
    )

def make_chart_title(text, sz=1200):
    run = RegularTextRun(rPr=make_jp_font(sz=sz, bold=True), t=text)
    para = Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=sz, bold=True)), r=[run])
    rich = RichText(p=[para])
    tx = Text(rich=rich)
    return Title(tx=tx)

from openpyxl.chart import LineChart, Reference

# メインの折れ線グラフ（左Y軸: 割合）
line = LineChart()
line.title = make_chart_title('④ 月別 解決率・任意退去率・解除件数の推移', sz=1100)
line.style = 10
line.width = 22
line.height = 14

# 解決率・任意退去率（B-C列）
data_pct = Reference(ws, min_col=2, min_row=22, max_col=3, max_row=29)
cats_ref = Reference(ws, min_col=1, min_row=23, max_row=29)
line.add_data(data_pct, titles_from_data=True)
line.set_categories(cats_ref)

# 左Y軸（割合）
line.y_axis.title = '割合'
line.y_axis.numFmt = '0%'
line.y_axis.scaling.min = 0
line.y_axis.scaling.max = 0.6
line.y_axis.majorUnit = 0.1
line.y_axis.delete = False
line.y_axis.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=900)), r=[])]
)

# X軸
line.x_axis.delete = False
line.x_axis.tickLblPos = 'low'
line.x_axis.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)), r=[])]
)

# 解決率: 青線
s0 = line.series[0]
s0.graphicalProperties.line.solidFill = '4472C4'
s0.graphicalProperties.line.width = 28000
s0.marker.symbol = 'circle'
s0.marker.size = 8
s0.marker.graphicalProperties.solidFill = '4472C4'
s0.dLbls = None

# 任意退去率: 緑線
s1 = line.series[1]
s1.graphicalProperties.line.solidFill = '92D050'
s1.graphicalProperties.line.width = 28000
s1.marker.symbol = 'square'
s1.marker.size = 8
s1.marker.graphicalProperties.solidFill = '92D050'
s1.dLbls = None

# ---- 第2軸: 解除件数（棒グラフ風に見せたいが、折れ線で統一） ----
# 第2の折れ線チャート（右Y軸: 件数）
line2 = LineChart()
data_cnt = Reference(ws, min_col=4, min_row=22, max_col=4, max_row=29)
line2.add_data(data_cnt, titles_from_data=True)
line2.set_categories(cats_ref)

# 右Y軸（件数）
line2.y_axis.title = '件数'
line2.y_axis.numFmt = '0'
line2.y_axis.scaling.min = 0
line2.y_axis.scaling.max = 140
line2.y_axis.majorUnit = 20
line2.y_axis.axId = 200  # 別軸ID
line2.y_axis.delete = False
line2.y_axis.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=900)), r=[])]
)

# 解除件数: オレンジ線（破線で差別化）
s2 = line2.series[0]
s2.graphicalProperties.line.solidFill = 'ED7D31'
s2.graphicalProperties.line.width = 28000
s2.graphicalProperties.line.dashStyle = 'dash'
s2.marker.symbol = 'diamond'
s2.marker.size = 8
s2.marker.graphicalProperties.solidFill = 'ED7D31'
s2.dLbls = None

# 第2チャートを第1チャートに合成（右Y軸に紐付け）
line2.y_axis.crosses = 'max'  # 右側に配置
line += line2

# 凡例
line.legend.position = 'b'
line.legend.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)), r=[])]
)

ws.add_chart(line, 'E62')

print('タスク②完了: 折れ線グラフに「解除件数」追加（オレンジ破線・右Y軸）')

# ============================================================
# 保存
# ============================================================
wb.save(FILE)
print('\n全タスク完了！')
