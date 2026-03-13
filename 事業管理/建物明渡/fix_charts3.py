# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText, Text
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties,
    Font as DrawingFont, RegularTextRun
)
from openpyxl.chart.title import Title
from openpyxl.styles import Font

FILE = r'C:\Users\user\Desktop\claudeマスター\事業管理\建物明渡\全保連解除＆退去案件まとめ【260309】.xlsx'

wb = load_workbook(FILE)

# グラフ分析シートを削除して作り直す
if 'グラフ分析' in wb.sheetnames:
    del wb['グラフ分析']

ws = wb.create_sheet('グラフ分析')

# ========== スタイル ==========
PIE_COLORS = ['92D050', '5B9BD5', 'BFBFBF']
FONT_HEADER = Font(name='游ゴシック', size=11, bold=True)
FONT_DATA = Font(name='游ゴシック', size=10)
FONT_TITLE = Font(name='游ゴシック', size=12, bold=True)

def make_jp_font(sz=1000, bold=False):
    return CharacterProperties(
        latin=DrawingFont(typeface='游ゴシック'),
        ea=DrawingFont(typeface='游ゴシック'),
        sz=sz, b=bold
    )

def make_chart_title(text, sz=1200):
    run = RegularTextRun(rPr=make_jp_font(sz=sz, bold=True), t=text)
    para = Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=sz, bold=True)), r=[run])
    rich = RichText(p=[para])
    tx = Text(rich=rich)
    return Title(tx=tx)

# ========== 合算シートからデータ再集計 ==========
ws_data = wb['25年7月～26年1月合算']

total = 0; taijo_total = 0; kaishuu_total = 0
pre_total = 0; pre_taijo = 0; pre_kaishuu = 0
post_total = 0; post_taijo = 0; post_kaishuu = 0

for row in range(13, ws_data.max_row + 1):
    if ws_data.cell(row=row, column=1).value is None:
        break
    total += 1
    g = ws_data.cell(row=row, column=7).value
    i = ws_data.cell(row=row, column=9).value
    j = ws_data.cell(row=row, column=10).value
    has_lawsuit = g is not None
    if has_lawsuit:
        post_total += 1
        if i is not None:
            post_taijo += 1; taijo_total += 1
        elif j is not None:
            post_kaishuu += 1; kaishuu_total += 1
    else:
        pre_total += 1
        if i is not None:
            pre_taijo += 1; taijo_total += 1
        elif j is not None:
            pre_kaishuu += 1; kaishuu_total += 1

mikaiketsu_total = total - taijo_total - kaishuu_total
pre_mikaiketsu = pre_total - pre_taijo - pre_kaishuu
post_mikaiketsu = post_total - post_taijo - post_kaishuu

# ========== 月別集計 ==========
month_sheets = ['25年7月解除','25年8月解除','25年9月解除','25年10月解除',
                '25年11月解除','25年12月解除','26年1月解除']
month_labels = ['7月','8月','9月','10月','11月','12月','1月']

monthly_data = []
for sn in month_sheets:
    ms = wb[sn]
    mt = 0; m_tai = 0; m_kai = 0
    for row in range(13, ms.max_row + 1):
        if ms.cell(row=row, column=1).value is None: break
        mt += 1
        iv = ms.cell(row=row, column=9).value
        jv = ms.cell(row=row, column=10).value
        if iv is not None: m_tai += 1
        elif jv is not None: m_kai += 1
    kr = (m_tai + m_kai) / mt if mt else 0
    tr = m_tai / mt if mt else 0
    monthly_data.append((kr, tr))

# ========== データ配置（A-C列、コンパクトに上部に） ==========

# タイトル
ws['A1'] = '【2025年7月～2026年1月 解除通知発送案件】グラフ分析（26年3月9日時点）'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

# --- ① 全体データ (rows 3-7) ---
ws['A3'] = '① 全体'; ws['A3'].font = FONT_HEADER
ws['A4'] = 'カテゴリ'; ws['B4'] = '件数'; ws['C4'] = '割合'
for c in ['A4','B4','C4']: ws[c].font = FONT_HEADER

pie1 = [('任意退去', taijo_total), ('全額入金（回収）', kaishuu_total), ('未解決', mikaiketsu_total)]
for idx, (cat, cnt) in enumerate(pie1):
    r = 5 + idx
    ws.cell(r, 1, cat).font = FONT_DATA
    ws.cell(r, 2, cnt).font = FONT_DATA
    c = ws.cell(r, 3, cnt / total); c.font = FONT_DATA; c.number_format = '0.0%'

# --- ② 提訴前データ (rows 9-13) ---
ws['A9'] = '② 提訴前'; ws['A9'].font = FONT_HEADER
ws['A10'] = 'カテゴリ'; ws['B10'] = '件数'; ws['C10'] = '割合'
for c in ['A10','B10','C10']: ws[c].font = FONT_HEADER

pie2 = [('任意退去', pre_taijo), ('全額入金（回収）', pre_kaishuu), ('未解決', pre_mikaiketsu)]
for idx, (cat, cnt) in enumerate(pie2):
    r = 11 + idx
    ws.cell(r, 1, cat).font = FONT_DATA
    ws.cell(r, 2, cnt).font = FONT_DATA
    c = ws.cell(r, 3, cnt / pre_total); c.font = FONT_DATA; c.number_format = '0.0%'

# --- ③ 提訴後データ (rows 15-19) ---
ws['A15'] = '③ 提訴後'; ws['A15'].font = FONT_HEADER
ws['A16'] = 'カテゴリ'; ws['B16'] = '件数'; ws['C16'] = '割合'
for c in ['A16','B16','C16']: ws[c].font = FONT_HEADER

pie3 = [('任意退去', post_taijo), ('全額入金（回収）', post_kaishuu), ('未解決', post_mikaiketsu)]
for idx, (cat, cnt) in enumerate(pie3):
    r = 17 + idx
    ws.cell(r, 1, cat).font = FONT_DATA
    ws.cell(r, 2, cnt).font = FONT_DATA
    c = ws.cell(r, 3, cnt / post_total); c.font = FONT_DATA; c.number_format = '0.0%'

# --- ④ 月別データ (rows 21-29) ---
ws['A21'] = '④ 月別推移'; ws['A21'].font = FONT_HEADER
ws['A22'] = '月'; ws['B22'] = '解決率'; ws['C22'] = '任意退去率'
for c in ['A22','B22','C22']: ws[c].font = FONT_HEADER

for idx, (kr, tr) in enumerate(monthly_data):
    r = 23 + idx
    ws.cell(r, 1, month_labels[idx]).font = FONT_DATA
    c1 = ws.cell(r, 2, kr); c1.font = FONT_DATA; c1.number_format = '0.0%'
    c2 = ws.cell(r, 3, tr); c2.font = FONT_DATA; c2.number_format = '0.0%'

# 列幅
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12

# ========== グラフ作成（十分な間隔で配置） ==========
# 1チャート ≈ 高さ13cm ≈ 25行分
# 配置: E列にチャート、各チャート間を十分に離す

def create_pie(ws, title_text, cat_rows, pct_col, anchor):
    pie = PieChart()
    pie.title = make_chart_title(title_text, sz=1100)
    pie.style = 10

    data_ref = Reference(ws, min_col=pct_col, min_row=cat_rows[0], max_row=cat_rows[-1])
    cats_ref = Reference(ws, min_col=1, min_row=cat_rows[0], max_row=cat_rows[-1])
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(cats_ref)

    series = pie.series[0]
    for i, color in enumerate(PIE_COLORS):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)

    series.dLbls = DataLabelList()
    series.dLbls.showPercent = False
    series.dLbls.showCatName = True
    series.dLbls.showVal = True
    series.dLbls.showSerName = False
    series.dLbls.showLeaderLines = True
    series.dLbls.numFmt = '0.0%'
    series.dLbls.separator = '\n'
    series.dLbls.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)), r=[])]
    )

    pie.legend = None
    pie.width = 14
    pie.height = 12
    ws.add_chart(pie, anchor)

# ① E2  (rows 2~20くらい)
create_pie(ws, f'① 全体の解決状況（全{total}件）', [5,6,7], 3, 'E2')

# ② E22 (rows 22~40くらい)
create_pie(ws, f'② 提訴前の解決状況（全{pre_total}件）', [11,12,13], 3, 'E22')

# ③ E42 (rows 42~60くらい)
create_pie(ws, f'③ 提訴後の解決状況（全{post_total}件）', [17,18,19], 3, 'E42')

# ④ 折れ線 E62 (rows 62~82くらい)
line = LineChart()
line.title = make_chart_title('④ 月別 解決率・任意退去率の推移', sz=1100)
line.style = 10
line.width = 18
line.height = 13

data_ref = Reference(ws, min_col=2, min_row=22, max_col=3, max_row=29)
cats_ref = Reference(ws, min_col=1, min_row=23, max_row=29)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)

# Y軸
line.y_axis.title = '割合'
line.y_axis.numFmt = '0%'
line.y_axis.scaling.min = 0
line.y_axis.scaling.max = 0.6
line.y_axis.majorUnit = 0.1
line.y_axis.delete = False
line.y_axis.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=900)), r=[])]
)

# X軸
line.x_axis.delete = False
line.x_axis.tickLblPos = 'low'
line.x_axis.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)), r=[])]
)

# 解決率: 青線
s0 = line.series[0]
s0.graphicalProperties.line.solidFill = '4472C4'
s0.graphicalProperties.line.width = 28000
s0.marker.symbol = 'circle'
s0.marker.size = 8
s0.marker.graphicalProperties.solidFill = '4472C4'
s0.dLbls = None

# 任意退去率: 緑線
s1 = line.series[1]
s1.graphicalProperties.line.solidFill = '92D050'
s1.graphicalProperties.line.width = 28000
s1.marker.symbol = 'square'
s1.marker.size = 8
s1.marker.graphicalProperties.solidFill = '92D050'
s1.dLbls = None

line.legend.position = 'b'
line.legend.txPr = RichText(
    p=[Paragraph(pPr=ParagraphProperties(defRPr=make_jp_font(sz=1000)), r=[])]
)

ws.add_chart(line, 'E62')

wb.save(FILE)
print('完了！チャートを重ならないように配置しました')
print(f'  ① 全体円グラフ: E2')
print(f'  ② 提訴前円グラフ: E22')
print(f'  ③ 提訴後円グラフ: E42')
print(f'  ④ 折れ線グラフ: E62')
