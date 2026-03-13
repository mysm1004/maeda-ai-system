import sys; sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.chart.layout import Layout, ManualLayout

src = '事業管理/建物明渡/【全保連様】25年7月から25年12月解除案件退去率まとめ【260310】.xlsx'
dst = '事業管理/建物明渡/【全保連様】25年7月から25年12月解除案件退去率まとめ【260310】_v2.xlsx'

wb = load_workbook(src)

# ============================================================
# 1. サマリー文言の修正
# ============================================================
def make_summary_text(prefix, taikyo_rate, kaiketsu_rate, nyukin_count):
    """実数値に基づいた自然なサマリー文を生成"""
    # 任意退去率の表現
    tr = taikyo_rate * 100
    if tr >= 45:
        taikyo_str = '約５割'
    elif tr >= 35:
        taikyo_str = '約４割'
    elif tr >= 25:
        taikyo_str = '約３割'
    elif tr >= 22:
        taikyo_str = '約２割半'
    elif tr >= 18:
        taikyo_str = '約２割'
    elif tr >= 12:
        taikyo_str = '約１割半'
    elif tr >= 8:
        taikyo_str = '約１割'
    else:
        taikyo_str = '１割未満'

    # 解決率(任意退去+全額入金)の表現
    kr = kaiketsu_rate * 100
    if kr >= 52:
        kaiketsu_str = '５０％超'
    elif kr >= 48:
        kaiketsu_str = '約５０％'
    elif kr >= 45:
        kaiketsu_str = '約４５％'
    elif kr >= 42:
        kaiketsu_str = '約４２％'
    elif kr >= 39:
        kaiketsu_str = '約４０％'
    elif kr >= 37:
        kaiketsu_str = '約３８％'
    elif kr >= 35:
        kaiketsu_str = '約３５％'
    else:
        kaiketsu_str = f'約{int(round(kr))}％'

    if nyukin_count > 0:
        return f'{prefix}サマリー：解除通知送付案件中、任意退去した割合が{taikyo_str}、全額入金も足すと{kaiketsu_str}'
    else:
        return f'{prefix}サマリー：解除通知送付案件中、任意退去した割合が{taikyo_str}'

# 各シートの実データ取得・サマリー更新
sheets_info = [
    ('25年7月～25年12月合算', None),
    ('25年7月解除', None), ('25年8月解除', None), ('25年9月解除', None),
    ('25年10月解除', None), ('25年11月解除', None), ('25年12月解除', None),
]

for name, _ in sheets_info:
    ws = wb[name]
    total = ws.cell(row=2, column=2).value
    kaiketsu = ws.cell(row=3, column=2).value
    taikyo = ws.cell(row=4, column=2).value
    nyukin = ws.cell(row=5, column=2).value
    taikyo_rate = taikyo / total if total else 0
    kaiketsu_rate = kaiketsu / total if total else 0

    old_text = ws.cell(row=1, column=1).value
    # Extract prefix (before サマリー)
    if 'サマリー' in old_text:
        prefix = old_text.split('サマリー')[0]
    else:
        prefix = old_text

    # 合算シートは少し表現を変える
    if '合算' in name:
        tr = taikyo_rate * 100
        kr = kaiketsu_rate * 100
        new_text = f'{prefix}サマリー：解除通知送付案件中、任意退去成功割合が約{int(round(tr))}％、全額入金も足すと約{int(round(kr))}％に到達'
    else:
        new_text = make_summary_text(prefix, taikyo_rate, kaiketsu_rate, nyukin)

    print(f'[{name}] taikyo={taikyo_rate:.1%} kaiketsu={kaiketsu_rate:.1%}')
    if old_text != new_text:
        print(f'  旧: ...{old_text.split("サマリー")[1] if "サマリー" in old_text else ""}')
        print(f'  新: ...{new_text.split("サマリー")[1]}')
        ws.cell(row=1, column=1).value = new_text
    else:
        print(f'  変更なし')

# ============================================================
# 2. ④折れ線グラフの右余白を詰める
# ============================================================
ws_g = wb['グラフ分析']
line_chart = ws_g._charts[3]  # Chart index 3 = ④

# Check current plot area layout
print(f'\n④グラフ plotArea layout: {line_chart.layout}')

# The chart itself is 15 x 7.5 (width x height in cm)
# With 7 data points it filled nicely; with 6 there's empty space on right
# Approach: adjust the chart's plotArea to use more of the available width
# Also can reduce chart width slightly

# Set plotArea to fill the chart better
# manualLayout with x, y, w, h as fractions of chart area
if line_chart.layout is None or line_chart.layout.manualLayout is None:
    line_chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.08, y=0.12,
            w=0.85, h=0.75,
            xMode='edge', yMode='edge',
            wMode='factor', hMode='factor'
        )
    )
else:
    ml = line_chart.layout.manualLayout
    print(f'  Current: x={ml.x}, y={ml.y}, w={ml.w}, h={ml.h}')
    # Expand width to fill the gap
    ml.w = 0.85
    ml.h = 0.75
    ml.x = 0.08
    ml.y = 0.12

# Also reduce the overall chart width slightly to tighten
# Original was 7920000 EMU wide (for 7 data points)
# For 6 points, scale proportionally: 7920000 * 6/7 ≈ 6788571
# But let's keep it a bit wider for readability
line_chart.width = 13.5  # cm (was 15)

print(f'  Chart width adjusted: 15 -> 13.5 cm')
print(f'  plotArea expanded to fill chart')

wb.save(dst)
print(f'\nSaved: {dst}')
